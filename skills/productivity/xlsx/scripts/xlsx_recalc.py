#!/usr/bin/env python3
"""Recalculate a workbook's formulas headlessly.

openpyxl never computes formulas. This script uses Excel COM automation
(Windows) to open, recalculate, and re-save the workbook, falling back to
`soffice` (LibreOffice) on non-Windows / no-Office environments. Cached
formula results then become available to `xlsx_read.py --data-only` and
`--formulas`.

Behavior:
  * Excel COM available: recalculates via Excel and writes --out (or
    replaces the input). Prints {"recalculated": true, ...} and exits 0.
  * Excel COM unavailable but soffice on PATH: converts the file to .xlsx
    in a temp dir (which recalculates all formulas) and writes --out.
  * Neither available: prints {"recalculated": false, "reason": ...} with
    installation guidance and STILL exits 0 — callers can branch on the
    JSON instead of the exit code.

Usage:
  xlsx_recalc.py book.xlsx
  xlsx_recalc.py book.xlsx --out recalced.xlsx
  xlsx_recalc.py book.xlsx --timeout 120
"""
from __future__ import annotations

import argparse
import json
import os
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path


def count_cached(path):
    """Number of formula cells with a cached value present."""
    from openpyxl import load_workbook
    wb_f = load_workbook(path, data_only=False)
    wb_v = load_workbook(path, data_only=True)
    formulas = cached = 0
    for name in wb_f.sheetnames:
        ws_f, ws_v = wb_f[name], wb_v[name]
        for row in ws_f.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas += 1
                    if ws_v[cell.coordinate].value is not None:
                        cached += 1
    return formulas, cached


def _find_repo_root() -> str | None:
    """Locate the repo root (the directory containing ``tools/``) from __file__."""
    current = os.path.dirname(os.path.realpath(__file__))
    for _ in range(8):
        if os.path.isfile(os.path.join(current, "tools", "office_excel_recalc.py")):
            return current
        parent = os.path.dirname(current)
        if parent == current:
            return None
        current = parent
    return None


def _recalc_via_com(src: Path, dest: Path) -> str | None:
    """Recalculate via Excel COM (no LibreOffice required).

    Returns the output path, or None when the shared converter is unavailable
    or fails — callers then fall back to soffice.
    """
    try:
        root = _find_repo_root()
        if root and root not in sys.path:
            sys.path.insert(0, root)
        from tools.office_excel_recalc import recalc_via_com
        out_path = str(dest) if dest != src else None
        return recalc_via_com(str(src), out_path=out_path)
    except Exception:
        return None


def main(argv=None):
    ap = argparse.ArgumentParser(
        description="Recalculate .xlsx formulas headlessly via Excel COM or LibreOffice.")
    ap.add_argument("file", help="path to .xlsx file")
    ap.add_argument("--out", help="output path (default: replace input)")
    ap.add_argument("--timeout", type=int, default=180,
                    help="seconds to wait for the recalc engine (default 180)")
    args = ap.parse_args(argv)

    src = Path(args.file).resolve()
    if not src.exists():
        print(json.dumps({"ok": False, "error": f"no such file: {src}"}),
              file=sys.stderr)
        return 1

    dest = Path(args.out).resolve() if args.out else src

    # 1. Prefer Excel COM (no LibreOffice dependency).
    output = _recalc_via_com(src, dest)
    if output is not None:
        formulas, cached = count_cached(output)
        print(json.dumps({
            "ok": True, "recalculated": True, "output": str(output),
            "formula_cells": formulas, "with_cached_values": cached,
        }, ensure_ascii=False))
        return 0

    # 2. Fall back to LibreOffice.
    soffice = shutil.which("soffice")
    if not soffice:
        print(json.dumps({
            "ok": True, "recalculated": False,
            "reason": "No recalc engine — LibreOffice (soffice) absent and "
                      "Excel COM unavailable",
            "guidance": "Install LibreOffice (e.g. `apt install "
                        "libreoffice-calc` or `brew install --cask "
                        "libreoffice`) or Microsoft Excel, or open the file "
                        "in Excel/LibreOffice once and re-save it.",
        }, ensure_ascii=False))
        return 0

    with tempfile.TemporaryDirectory() as tmp:
        proc = subprocess.run(
            [soffice, "--headless", "--calc", "--convert-to", "xlsx:Calc "
             "MS Excel 2007 XML", "--outdir", tmp, str(src)],
            capture_output=True, text=True, encoding="utf-8",
            timeout=args.timeout,
            env={"HOME": tmp, "PATH": Path(soffice).parent.as_posix()
                 + ":/usr/bin:/bin"})
        produced = Path(tmp) / (src.stem + ".xlsx")
        if proc.returncode != 0 or not produced.exists():
            print(json.dumps({"ok": False,
                              "error": "soffice conversion failed",
                              "stderr": proc.stderr.strip()[-500:]}),
                  file=sys.stderr)
            return 1
        formulas, cached = count_cached(produced)
        shutil.copyfile(produced, dest)

    print(json.dumps({
        "ok": True, "recalculated": True, "output": str(dest),
        "formula_cells": formulas, "with_cached_values": cached,
    }, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except Exception as exc:  # noqa: BLE001
        print(json.dumps({"ok": False, "error": str(exc)}), file=sys.stderr)
        sys.exit(1)
