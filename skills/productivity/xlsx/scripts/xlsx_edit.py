#!/usr/bin/env python3
"""Edit an existing .xlsx workbook in place (or to --out).

Operations (repeatable where noted, applied in the order listed below):
  --rename-sheet OLD:NEW        rename a sheet
  --copy-sheet SRC:NEW          duplicate a sheet under a new name
  --insert-rows IDX[:N]         insert N rows before row IDX (default N=1)
  --delete-rows IDX[:N]         delete N rows starting at row IDX
  --insert-cols IDX[:N]         insert N columns before column IDX (number)
  --delete-cols IDX[:N]         delete N columns starting at column IDX
  --set CELL=VALUE              repeatable; type-inferred (int, float, bool,
                                ISO date, else string). '=...' sets a formula.
  --append ROWJSON              repeatable; JSON array appended as a row
  --recalc                      set fullCalcOnLoad so Excel/LibreOffice
                                recomputes all formulas on next open

WARNING: openpyxl does NOT shift merged-cell ranges, chart anchors, or
formula references when rows/columns are inserted or deleted. Verify any
sheet containing merges or formulas after structural edits.

Usage:
  xlsx_edit.py book.xlsx --sheet Data --set B2=42 --set C2=2026-01-01 \
      --set "D2==SUM(B2:C2)" --recalc
  xlsx_edit.py book.xlsx --sheet Data --append '["Widget", 9.99, true]'
  xlsx_edit.py book.xlsx --copy-sheet Data:Backup --rename-sheet Data:Main
"""
from __future__ import annotations

import argparse
import json
import sys
from datetime import date, datetime

from openpyxl import load_workbook


def infer(text):
    if text.startswith("="):
        return text  # formula
    low = text.lower()
    if low in ("true", "false"):
        return low == "true"
    for caster in (int, float):
        try:
            return caster(text)
        except ValueError:
            pass
    for parser in (date.fromisoformat, datetime.fromisoformat):
        try:
            return parser(text)
        except ValueError:
            pass
    return text


def parse_idx(arg):
    if ":" in arg:
        idx, n = arg.split(":", 1)
        return int(idx), int(n)
    return int(arg), 1


def main(argv=None):
    ap = argparse.ArgumentParser(
        description="Edit an existing .xlsx workbook.",
        epilog="openpyxl does not shift merges/formula refs on insert/delete.")
    ap.add_argument("file", help="path to .xlsx file")
    ap.add_argument("--sheet", help="target sheet (default: active)")
    ap.add_argument("--out", help="output path (default: edit in place)")
    ap.add_argument("--rename-sheet", action="append", default=[],
                    metavar="OLD:NEW")
    ap.add_argument("--copy-sheet", action="append", default=[],
                    metavar="SRC:NEW")
    ap.add_argument("--insert-rows", action="append", default=[],
                    metavar="IDX[:N]")
    ap.add_argument("--delete-rows", action="append", default=[],
                    metavar="IDX[:N]")
    ap.add_argument("--insert-cols", action="append", default=[],
                    metavar="IDX[:N]")
    ap.add_argument("--delete-cols", action="append", default=[],
                    metavar="IDX[:N]")
    ap.add_argument("--set", action="append", default=[], metavar="CELL=VALUE")
    ap.add_argument("--append", action="append", default=[], metavar="ROWJSON")
    ap.add_argument("--recalc", action="store_true",
                    help="force full recalculation when the file is opened")
    args = ap.parse_args(argv)

    wb = load_workbook(args.file)
    changes = []

    for pair in args.rename_sheet:
        old, new = pair.split(":", 1)
        wb[old].title = new
        changes.append(f"rename {old}->{new}")
    for pair in args.copy_sheet:
        src, new = pair.split(":", 1)
        copy = wb.copy_worksheet(wb[src])
        copy.title = new
        changes.append(f"copy {src}->{new}")

    ws = wb[args.sheet] if args.sheet else wb.active

    for arg in args.insert_rows:
        idx, n = parse_idx(arg)
        ws.insert_rows(idx, n)
        changes.append(f"insert_rows {idx}x{n}")
    for arg in args.delete_rows:
        idx, n = parse_idx(arg)
        ws.delete_rows(idx, n)
        changes.append(f"delete_rows {idx}x{n}")
    for arg in args.insert_cols:
        idx, n = parse_idx(arg)
        ws.insert_cols(idx, n)
        changes.append(f"insert_cols {idx}x{n}")
    for arg in args.delete_cols:
        idx, n = parse_idx(arg)
        ws.delete_cols(idx, n)
        changes.append(f"delete_cols {idx}x{n}")

    for assignment in args.set:
        coord, raw = assignment.split("=", 1)
        ws[coord] = infer(raw)
        changes.append(f"set {coord}")
    for row_json in args.append:
        ws.append(json.loads(row_json))
        changes.append(f"append row {ws.max_row}")

    if args.recalc:
        wb.calculation.fullCalcOnLoad = True
        changes.append("fullCalcOnLoad")

    out = args.out or args.file
    wb.save(out)
    print(json.dumps({"ok": True, "output": out, "sheet": ws.title,
                      "changes": changes}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except Exception as exc:  # noqa: BLE001
        print(json.dumps({"ok": False, "error": str(exc)}), file=sys.stderr)
        sys.exit(1)
