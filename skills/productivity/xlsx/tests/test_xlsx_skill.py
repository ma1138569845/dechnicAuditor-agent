"""End-to-end tests for the xlsx skill helper scripts.

Runs each script as a subprocess under LC_ALL=C to prove all text I/O
uses explicit UTF-8 rather than locale defaults. No network access.
"""
from __future__ import annotations

import csv
import json
import os
import subprocess
import sys
from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook

SCRIPTS = Path(__file__).resolve().parent.parent / "scripts"


def run(script, *args, expect_ok=True):
    env = dict(os.environ, LC_ALL="C", LANG="C")
    env.pop("PYTHONIOENCODING", None)
    proc = subprocess.run(
        [sys.executable, str(SCRIPTS / script), *map(str, args)],
        capture_output=True, text=True, env=env, encoding="utf-8")
    if expect_ok:
        assert proc.returncode == 0, f"{script} failed: {proc.stderr}"
    return proc


SPEC = {
    "full_calc_on_load": True,
    "sheets": [
        {
            "name": "Data",
            "rows": [
                [
                    {"value": "Region", "bold": True, "fill": "DDEBF7",
                     "border": "thin", "align": "center", "valign": "center"},
                    {"value": "Sales", "bold": True, "fill": "DDEBF7"},
                    {"value": "Growth", "bold": True},
                    {"value": "Audited", "bold": True},
                    {"value": "Closed", "bold": True},
                    {"value": "Status", "bold": True},
                ],
                ["North", 1500.5, {"value": 0.125, "format": "0.0%"}, True,
                 {"value": "2026-01-31", "type": "date",
                  "format": "yyyy-mm-dd"}, "Yes"],
                ["South", 900, {"value": -0.03, "format": "0.0%"}, False,
                 {"value": "2026-02-28", "type": "date",
                  "format": "yyyy-mm-dd"}, "No"],
                ["East", 2100, {"value": 0.4, "format": "0.0%"}, True,
                 {"value": "2026-03-31", "type": "date",
                  "format": "yyyy-mm-dd"}, "Yes"],
            ],
            "cells": {
                "A6": {"value": "Total", "bold": True, "italic": True,
                       "font_size": 12, "font_color": "1F4E78"},
                "B6": {"formula": "SUM(B2:B4)", "format": "$#,##0.00"},
            },
            "column_widths": {"A": 18, "B": 14},
            "row_heights": {"1": 24},
            "merges": ["A8:C8"],
            "freeze_panes": "A2",
            "autofilter": "A1:F4",
            "conditional_formats": [
                {"range": "B2:B4", "type": "cell_is",
                 "operator": "greaterThan", "formula": ["1000"],
                 "fill": "C6EFCE"},
                {"range": "C2:C4", "type": "color_scale"},
            ],
            "charts": [
                {"type": "bar", "title": "Sales by region", "anchor": "H2",
                 "data": "B1:B4", "categories": "A2:A4"},
                {"type": "line", "title": "Growth", "anchor": "H18",
                 "data": "C1:C4", "categories": "A2:A4"},
                {"type": "pie", "title": "Share", "anchor": "P2",
                 "data": "B2:B4", "categories": "A2:A4",
                 "titles_from_data": False},
            ],
            "validations": [
                {"range": "F2:F10", "type": "list",
                 "formula1": '"Yes,No,Maybe"'},
            ],
        },
        {"name": "Notes", "rows": [["Zürich", "Фамилия", "12,5%"]]},
    ],
}


@pytest.fixture
def workbook(tmp_path):
    spec_path = tmp_path / "spec.json"
    spec_path.write_text(json.dumps(SPEC), encoding="utf-8")
    out = tmp_path / "report.xlsx"
    proc = run("xlsx_create.py", spec_path, out)
    summary = json.loads(proc.stdout)
    assert summary["ok"] and summary["sheets"] == ["Data", "Notes"]
    return out


def test_create_features_roundtrip(workbook):
    wb = load_workbook(workbook)
    ws = wb["Data"]
    # typed values
    assert ws["B2"].value == 1500.5
    assert ws["D2"].value is True
    e2 = ws["E2"].value
    assert (e2.date() if hasattr(e2, "date") else e2) == date(2026, 1, 31)
    # formula + number formats
    assert ws["B6"].value == "=SUM(B2:B4)"
    assert ws["B6"].number_format == "$#,##0.00"
    assert ws["C2"].number_format == "0.0%"
    assert ws["E2"].number_format == "yyyy-mm-dd"
    # styling
    assert ws["A1"].font.bold is True
    assert ws["A1"].fill.fgColor.rgb.endswith("DDEBF7")
    assert ws["A1"].border.left.style == "thin"
    assert ws["A1"].alignment.horizontal == "center"
    assert ws["A6"].font.italic is True and ws["A6"].font.size == 12
    # dimensions
    assert ws.column_dimensions["A"].width == 18
    assert ws.row_dimensions[1].height == 24
    # merges / freeze / autofilter
    assert "A8:C8" in [str(r) for r in ws.merged_cells.ranges]
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref == "A1:F4"
    # conditional formatting, charts, validation
    assert len(list(ws.conditional_formatting)) == 2
    assert len(ws._charts) == 3
    types = {type(c).__name__ for c in ws._charts}
    assert types == {"BarChart", "LineChart", "PieChart"}
    assert len(ws.data_validations.dataValidation) == 1
    # recalc flag
    assert wb.calculation.fullCalcOnLoad is True


def test_read_sheets_json_formulas(workbook, tmp_path):
    inv = json.loads(run("xlsx_read.py", workbook, "--sheets").stdout)
    names = [s["name"] for s in inv["sheets"]]
    assert names == ["Data", "Notes"]
    data_info = inv["sheets"][0]
    assert data_info["charts"] == 3
    assert "A8:C8" in data_info["merged"]
    assert data_info["freeze_panes"] == "A2"

    dump = json.loads(
        run("xlsx_read.py", workbook, "--json", "--sheet", "Data").stdout)
    assert dump["rows"][1][0] == "North"
    assert dump["rows"][1][4] == "2026-01-31T00:00:00"

    notes = json.loads(
        run("xlsx_read.py", workbook, "--json", "--sheet", "Notes").stdout)
    assert notes["rows"][0] == ["Zürich", "Фамилия", "12,5%"]

    formulas = json.loads(run("xlsx_read.py", workbook, "--formulas").stdout)
    entry = [f for f in formulas["formulas"] if f["cell"] == "B6"][0]
    assert entry["formula"] == "=SUM(B2:B4)"
    # openpyxl never computes: cached value absent on a fresh file
    assert entry["cached"] is None

    csv_out = tmp_path / "data.csv"
    run("xlsx_read.py", workbook, "--csv", "--sheet", "Notes",
        "--out", csv_out)
    text = csv_out.read_text(encoding="utf-8")
    assert "Zürich" in text and "Фамилия" in text


def test_csv_roundtrip_nonascii(tmp_path):
    src = tmp_path / "src.csv"
    with open(src, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["City", "Share", "Surname", "Active", "When"])
        w.writerow(["Zürich", "12,5%", "Фамилия", "true", "2026-05-01"])
        w.writerow(["Oslo", "7", "Ås", "false", "2026-06-01"])
    xlsx = tmp_path / "conv.xlsx"
    run("csv_to_xlsx.py", src, xlsx, "--sheet-name", "Import")

    wb = load_workbook(xlsx)
    ws = wb["Import"]
    assert ws["A2"].value == "Zürich"
    assert ws["B2"].value == "12,5%"      # decimal comma stays a string
    assert ws["C2"].value == "Фамилия"
    assert ws["D2"].value is True          # bool inferred
    assert ws["E2"].value.date() == date(2026, 5, 1)  # date inferred
    assert ws["B3"].value == 7             # int inferred
    assert ws["A1"].font.bold is True      # styled header
    assert ws.freeze_panes == "A2"

    back = tmp_path / "back.csv"
    run("xlsx_to_csv.py", xlsx, back, "--sheet", "Import")
    with open(back, newline="", encoding="utf-8") as fh:
        rows = list(csv.reader(fh))
    assert rows[1][0] == "Zürich"
    assert rows[1][2] == "Фамилия"
    assert rows[1][3] == "True"
    assert rows[1][4] == "2026-05-01"

    # encoding override
    latin = tmp_path / "latin.csv"
    run("xlsx_to_csv.py", xlsx, latin, "--sheet", "Import",
        "--encoding", "utf-8-sig")
    assert latin.read_bytes().startswith(b"\xef\xbb\xbf")


def test_edit_existing(workbook, tmp_path):
    edited = tmp_path / "edited.xlsx"
    proc = run("xlsx_edit.py", workbook, "--sheet", "Notes",
               "--out", edited,
               "--copy-sheet", "Notes:Backup",
               "--rename-sheet", "Data:Main",
               "--set", "B1=Änderung",
               "--set", "C1=99.5",
               "--set", "D1=2026-12-24",
               "--set", "E1==SUM(C1:C1)",
               "--append", '["appended", 1, false]',
               "--insert-rows", "1:1",
               "--recalc")
    result = json.loads(proc.stdout)
    assert result["ok"]

    wb = load_workbook(edited)
    assert set(wb.sheetnames) == {"Main", "Notes", "Backup"}
    ws = wb["Notes"]
    # insert-rows ran before --set per documented order, so row 1 is blank
    # and original data moved to row 2... check documented ordering:
    # structural ops run before --set, so B1 etc. were written after insert.
    assert ws["B1"].value == "Änderung"
    assert ws["C1"].value == 99.5
    assert ws["D1"].value.date() == date(2026, 12, 24)
    assert ws["E1"].value == "=SUM(C1:C1)"
    assert wb.calculation.fullCalcOnLoad is True
    # appended row present
    found = [r for r in ws.iter_rows(values_only=True)
             if r and r[0] == "appended"]
    assert found and found[0][1] == 1 and found[0][2] is False
    # copy preserved data
    assert wb["Backup"]["A1"].value == "Zürich"


def test_help_and_errors():
    for script in ["xlsx_create.py", "xlsx_read.py", "xlsx_edit.py",
                   "csv_to_xlsx.py", "xlsx_to_csv.py"]:
        proc = run(script, "--help")
        assert "usage" in proc.stdout.lower()
    bad = run("xlsx_read.py", "/nonexistent.xlsx", "--sheets",
              expect_ok=False)
    assert bad.returncode != 0
    assert json.loads(bad.stderr)["ok"] is False
